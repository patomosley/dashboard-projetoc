"""
Script para criar um arquivo Excel de exemplo para importação
Execute este script para gerar um arquivo de exemplo
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Projetos"

# Definir estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Cabeçalhos
headers = ['ID', 'Protocolo', 'Nome', 'Valor Mensal', 'Contato', 'Data Agendamento', 'Tipo Cliente', 'Status', 'Data Entrega']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

# Dados de exemplo
example_data = [
    ['', 'PROT001', 'Projeto Website', 5000.00, 'João Silva', '15/02/2026', 'B2G', 'Pendente', ''],
    ['', 'PROT002', 'App Mobile', 8500.50, 'Maria Santos', '20/02/2026', 'ISP', 'Em Andamento', ''],
    ['', 'PROT003', 'Consultoria TI', 3200.00, 'Carlos Oliveira', '10/02/2026', 'B2B', 'Concluído', '28/02/2026'],
    ['', 'PROT004', 'Suporte Técnico', 2000.00, 'Ana Costa', '25/02/2026', 'B2G', 'Pendente', ''],
    ['', 'PROT005', 'Infraestrutura', 6500.75, 'Pedro Alves', '05/02/2026', 'ISP', 'Atrasado', ''],
]

# Adicionar dados
for row_num, row_data in enumerate(example_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border
        if col_num == 4:  # Coluna de valor
            cell.number_format = '#,##0.00'

# Ajustar largura das colunas
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 18

# Salvar arquivo
wb.save('exemplo_projetos.xlsx')
print("Arquivo 'exemplo_projetos.xlsx' criado com sucesso!")
