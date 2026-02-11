from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from models import db, Project, Observation, Technology, Service
from datetime import datetime, timedelta
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv

load_dotenv()

# Configurações de e-mail (use variáveis de ambiente)
SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', 587))
EMAIL_USER = os.getenv('EMAIL_USER', 'seu_email@gmail.com')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD', 'sua_senha_app')

def send_email(recipient_email, subject, body, html_body=None):
    """Envia e-mail para o cliente"""
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = EMAIL_USER
        msg['To'] = recipient_email
        
        # Versão em texto plano
        msg.attach(MIMEText(body, 'plain'))
        
        # Versão em HTML (se fornecida)
        if html_body:
            msg.attach(MIMEText(html_body, 'html'))
        
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASSWORD)
            server.send_message(msg)
        
        return True, "E-mail enviado com sucesso!"
    except Exception as e:
        return False, f"Erro ao enviar e-mail: {str(e)}"

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///projects.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

with app.app_context():
    db.create_all()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/projects', methods=['GET'])
def get_projects():
    # Parâmetros de filtro
    status_filter = request.args.get('status', '').strip()
    client_type_filter = request.args.get('client_type', '').strip()
    search_query = request.args.get('search', '').strip()
    
    # Iniciar query
    query = Project.query
    
    # Aplicar filtros
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    if client_type_filter:
        query = query.filter_by(client_type=client_type_filter)
    
    if search_query:
        query = query.filter(
            (Project.name.ilike(f'%{search_query}%')) |
            (Project.contract_protocol.ilike(f'%{search_query}%')) |
            (Project.contact.ilike(f'%{search_query}%'))
        )
    
    projects = query.all()
    
    return jsonify([{
        'id': p.id,
        'contract_protocol': p.contract_protocol,
        'name': p.name,
        'monthly_value': p.monthly_value,
        'contact': p.contact,
        'scheduled_date': p.scheduled_date.strftime('%Y-%m-%d'),
        'client_type': p.client_type,
        'status': p.status,
        'delivery_date': p.delivery_date.strftime('%Y-%m-%d') if p.delivery_date else None
    } for p in projects])

@app.route('/api/projects', methods=['POST'])
def create_project():
    data = request.json
    new_project = Project(
        contract_protocol=data['contract_protocol'],
        name=data['name'],
        monthly_value=float(data['monthly_value']),
        contact=data['contact'],
        scheduled_date=datetime.strptime(data['scheduled_date'], '%Y-%m-%d'),
        client_type=data['client_type'],
        status='Pendente'
    )
    db.session.add(new_project)
    db.session.commit()
    return jsonify({'message': 'Projeto criado com sucesso!', 'id': new_project.id}), 201

@app.route('/api/projects/<int:id>', methods=['GET'])
def get_project(id):
    project = Project.query.get_or_404(id)
    return jsonify({
        'id': project.id,
        'contract_protocol': project.contract_protocol,
        'name': project.name,
        'monthly_value': project.monthly_value,
        'contact': project.contact,
        'scheduled_date': project.scheduled_date.strftime('%Y-%m-%d'),
        'client_type': project.client_type,
        'status': project.status,
        'delivery_date': project.delivery_date.strftime('%Y-%m-%d') if project.delivery_date else None
    })

@app.route('/api/projects/<int:id>', methods=['PUT'])
def update_project(id):
    project = Project.query.get_or_404(id)
    data = request.json
    
    project.contract_protocol = data.get('contract_protocol', project.contract_protocol)
    project.name = data.get('name', project.name)
    project.monthly_value = float(data.get('monthly_value', project.monthly_value))
    project.contact = data.get('contact', project.contact)
    project.scheduled_date = datetime.strptime(data.get('scheduled_date', project.scheduled_date.strftime('%Y-%m-%d')), '%Y-%m-%d')
    project.client_type = data.get('client_type', project.client_type)
    project.status = data.get('status', project.status)
    
    if data.get('status') == 'Concluído' and not project.delivery_date:
        project.delivery_date = datetime.utcnow()
    
    db.session.commit()
    return jsonify({'message': 'Projeto atualizado com sucesso!'}), 200

@app.route('/api/projects/<int:id>', methods=['DELETE'])
def delete_project(id):
    project = Project.query.get_or_404(id)
    db.session.delete(project)
    db.session.commit()
    return jsonify({'message': 'Projeto deletado com sucesso!'}), 200

@app.route('/api/projects/<int:id>/status', methods=['PATCH'])
def update_status(id):
    data = request.json
    project = Project.query.get_or_404(id)
    project.status = data['status']
    if data['status'] == 'Concluído':
        project.delivery_date = datetime.utcnow()
    db.session.commit()
    return jsonify({'message': 'Status atualizado!'})

@app.route('/api/projects/<int:id>/observations', methods=['GET', 'POST'])
def handle_observations(id):
    if request.method == 'POST':
        data = request.json
        obs = Observation(content=data['content'], project_id=id)
        db.session.add(obs)
        db.session.commit()
        return jsonify({'message': 'Observação adicionada!'})
    
    observations = Observation.query.filter_by(project_id=id).order_by(Observation.timestamp.desc()).all()
    return jsonify([{
        'content': o.content,
        'timestamp': o.timestamp.strftime('%d/%m/%Y %H:%M:%S')
    } for o in observations])

@app.route('/api/projects/<int:id>/send-email', methods=['POST'])
def send_project_email(id):
    """Envia e-mail para o cliente sobre o projeto"""
    project = Project.query.get_or_404(id)
    data = request.json
    
    subject = data.get('subject', f'Atualização - Projeto {project.name}')
    message = data.get('message', '')
    
    # Criar corpo do e-mail em HTML
    html_body = f"""
    <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px; background-color: #f8f9fa; border-radius: 8px;">
                <h2 style="color: #0d6efd; border-bottom: 2px solid #0d6efd; padding-bottom: 10px;">
                    Atualização de Projeto
                </h2>
                
                <div style="margin: 20px 0; background-color: white; padding: 15px; border-radius: 5px;">
                    <p><strong>Projeto:</strong> {project.name}</p>
                    <p><strong>Protocolo:</strong> {project.contract_protocol}</p>
                    <p><strong>Status:</strong> <span style="background-color: #e7f3ff; padding: 5px 10px; border-radius: 3px;">{project.status}</span></p>
                    <p><strong>Valor Mensal:</strong> R$ {project.monthly_value:,.2f}</p>
                    <p><strong>Data de Agendamento:</strong> {project.scheduled_date.strftime('%d/%m/%Y')}</p>
                </div>
                
                <div style="margin: 20px 0; background-color: white; padding: 15px; border-radius: 5px;">
                    <h3 style="color: #0d6efd; margin-top: 0;">Mensagem:</h3>
                    <p>{message}</p>
                </div>
                
                <div style="margin-top: 20px; padding-top: 20px; border-top: 1px solid #ddd; font-size: 12px; color: #666;">
                    <p>Este é um e-mail automático do sistema ProjectTracker. Não responda este e-mail.</p>
                </div>
            </div>
        </body>
    </html>
    """
    
    text_body = f"""
Atualização de Projeto

Projeto: {project.name}
Protocolo: {project.contract_protocol}
Status: {project.status}
Valor Mensal: R$ {project.monthly_value:,.2f}
Data de Agendamento: {project.scheduled_date.strftime('%d/%m/%Y')}

Mensagem:
{message}

---
Este é um e-mail automático do sistema ProjectTracker.
    """
    
    success, msg = send_email(project.contact, subject, text_body, html_body)
    
    if success:
        return jsonify({'message': msg}), 200
    else:
        return jsonify({'error': msg}), 400

@app.route('/api/alerts', methods=['GET'])
def get_alerts():
    """Retorna alertas inteligentes sobre projetos críticos"""
    projects = Project.query.all()
    alerts = []
    today = datetime.now().date()
    
    for project in projects:
        # Ignorar projetos cancelados ou concluídos
        if project.status in ['Cancelado', 'Concluído']:
            continue
        
        alert_data = {
            'id': project.id,
            'project_name': project.name,
            'type': None,
            'message': None,
            'severity': None,
            'action': None
        }
        
        # Alerta 1: Projetos atrasados
        if project.status == 'Atrasado':
            alert_data['type'] = 'atrasado'
            alert_data['severity'] = 'danger'
            alert_data['message'] = f'Projeto "{project.name}" está ATRASADO!'
            alert_data['action'] = 'Contatar cliente'
            alerts.append(alert_data)
        
        # Alerta 2: Projetos próximos do vencimento
        scheduled_date = project.scheduled_date.date()
        days_until = (scheduled_date - today).days
        
        if 0 <= days_until <= 3 and project.status in ['Pendente', 'Em Andamento']:
            alert_data['type'] = 'proximo_vencimento'
            alert_data['severity'] = 'warning'
            alert_data['message'] = f'Projeto "{project.name}" vence em {days_until} dia(s)!'
            alert_data['action'] = 'Acelerar entrega'
            alerts.append(alert_data)
        
        # Alerta 3: Projetos vencidos
        if scheduled_date < today and project.status not in ['Concluído', 'Cancelado']:
            alert_data['type'] = 'vencido'
            alert_data['severity'] = 'danger'
            alert_data['message'] = f'Projeto "{project.name}" está vencido há {(today - scheduled_date).days} dia(s)!'
            alert_data['action'] = 'Ação urgente'
            alerts.append(alert_data)
        
        # Alerta 4: Projetos em andamento há muito tempo
        if project.status == 'Em Andamento':
            days_in_progress = (today - project.scheduled_date.date()).days
            if days_in_progress > 30:
                alert_data['type'] = 'em_andamento_longo'
                alert_data['severity'] = 'info'
                alert_data['message'] = f'Projeto "{project.name}" está em andamento há {days_in_progress} dias'
                alert_data['action'] = 'Verificar progresso'
                alerts.append(alert_data)
    
    return jsonify(alerts)

@app.route('/api/stats')
def get_stats():
    # Parâmetros de filtro
    status_filter = request.args.get('status', '').strip()
    client_type_filter = request.args.get('client_type', '').strip()
    search_query = request.args.get('search', '').strip()
    
    # Iniciar query
    query = Project.query
    
    # Aplicar filtros
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    if client_type_filter:
        query = query.filter_by(client_type=client_type_filter)
    
    if search_query:
        query = query.filter(
            (Project.name.ilike(f'%{search_query}%')) |
            (Project.contract_protocol.ilike(f'%{search_query}%')) |
            (Project.contact.ilike(f'%{search_query}%'))
        )
    
    projects = query.all()
    total_revenue = sum(p.monthly_value for p in projects)
    
    # Estatísticas por mês (para gráfico de receita/entregas)
    monthly_stats = {}
    for p in projects:
        month = p.scheduled_date.strftime('%Y-%m')
        if month not in monthly_stats:
            monthly_stats[month] = {'revenue': 0, 'deliveries': 0}
        monthly_stats[month]['revenue'] += p.monthly_value
        if p.status == 'Concluído':
            monthly_stats[month]['deliveries'] += 1

    stats = {
        'total_projects': len(projects),
        'total_revenue': total_revenue,
        'status_counts': {
            'Pendente': Project.query.filter_by(status='Pendente').count(),
            'Em Andamento': Project.query.filter_by(status='Em Andamento').count(),
            'Concluído': Project.query.filter_by(status='Concluído').count(),
            'Atrasado': Project.query.filter_by(status='Atrasado').count(),
            'Cancelado': Project.query.filter_by(status='Cancelado').count()
        },
        'client_types': {
            'B2G': Project.query.filter_by(client_type='B2G').count(),
            'ISP': Project.query.filter_by(client_type='ISP').count(),
            'B2B': Project.query.filter_by(client_type='B2B').count()
        },
        'monthly_history': monthly_stats
    }
    return jsonify(stats)


@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """Exporta todos os projetos para um arquivo Excel"""
    projects = Project.query.all()
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Projetos"
    
    # Definir estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos
    headers = ['ID', 'Protocolo', 'Nome', 'Valor Mensal', 'Contato', 'Data Agendamento', 'Tipo Cliente', 'Status', 'Data Entrega']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Dados
    for row_num, project in enumerate(projects, 2):
        ws.cell(row=row_num, column=1).value = project.id
        ws.cell(row=row_num, column=2).value = project.contract_protocol
        ws.cell(row=row_num, column=3).value = project.name
        ws.cell(row=row_num, column=4).value = project.monthly_value
        ws.cell(row=row_num, column=5).value = project.contact
        ws.cell(row=row_num, column=6).value = project.scheduled_date.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=7).value = project.client_type
        ws.cell(row=row_num, column=8).value = project.status
        ws.cell(row=row_num, column=9).value = project.delivery_date.strftime('%d/%m/%Y') if project.delivery_date else ''
        
        # Aplicar bordas e formatação
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            if col_num == 4:  # Coluna de valor
                cell.number_format = '#,##0.00'
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    
    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'projetos_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/api/import/excel', methods=['POST'])
def import_excel():
    """Importa projetos de um arquivo Excel"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Arquivo não selecionado'}), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Apenas arquivos .xlsx são aceitos'}), 400
    
    try:
        from openpyxl import load_workbook
        
        # Carregar workbook
        wb = load_workbook(file.stream)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        # Iterar sobre as linhas (começando da linha 2, pulando cabeçalho)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Extrair dados (ajustar índices conforme necessário)
                protocol = row[1]  # Protocolo
                name = row[2]      # Nome
                value = row[3]     # Valor
                contact = row[4]   # Contato
                scheduled_date_str = row[5]  # Data Agendamento
                client_type = row[6]  # Tipo Cliente
                status = row[7] if row[7] else 'Pendente'  # Status
                
                # Validar dados obrigatórios
                if not all([protocol, name, value, contact, scheduled_date_str, client_type]):
                    errors.append(f"Linha {row_num}: Dados incompletos")
                    continue
                
                # Verificar se protocolo já existe
                if Project.query.filter_by(contract_protocol=protocol).first():
                    errors.append(f"Linha {row_num}: Protocolo '{protocol}' já existe")
                    continue
                
                # Converter data
                if isinstance(scheduled_date_str, str):
                    scheduled_date = datetime.strptime(scheduled_date_str, '%d/%m/%Y')
                else:
                    scheduled_date = scheduled_date_str
                
                # Criar projeto
                new_project = Project(
                    contract_protocol=protocol,
                    name=name,
                    monthly_value=float(value),
                    contact=contact,
                    scheduled_date=scheduled_date,
                    client_type=client_type,
                    status=status
                )
                
                db.session.add(new_project)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Linha {row_num}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'message': f'{imported_count} projetos importados com sucesso!',
            'imported_count': imported_count,
            'errors': errors
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Erro ao processar arquivo: {str(e)}'}), 400

# ===== ENDPOINTS DE CONFIGURAÇÕES (TECNOLOGIAS E SERVIÇOS) =====

@app.route('/api/technologies', methods=['GET'])
def get_technologies():
    """Retorna lista de tecnologias de entrega"""
    technologies = Technology.query.order_by(Technology.name).all()
    return jsonify([t.to_dict() for t in technologies])

@app.route('/api/technologies', methods=['POST'])
def create_technology():
    """Cria uma nova tecnologia de entrega"""
    data = request.get_json()
    name = data.get('name', '').strip()
    
    if not name:
        return jsonify({'error': 'Nome da tecnologia é obrigatório'}), 400
    
    # Verificar se já existe
    if Technology.query.filter_by(name=name).first():
        return jsonify({'error': 'Tecnologia já existe'}), 400
    
    tech = Technology(name=name)
    db.session.add(tech)
    db.session.commit()
    
    return jsonify(tech.to_dict()), 201

@app.route('/api/technologies/<int:tech_id>', methods=['DELETE'])
def delete_technology(tech_id):
    """Deleta uma tecnologia de entrega"""
    tech = Technology.query.get(tech_id)
    if not tech:
        return jsonify({'error': 'Tecnologia não encontrada'}), 404
    
    db.session.delete(tech)
    db.session.commit()
    
    return jsonify({'message': 'Tecnologia deletada com sucesso'}), 200

@app.route('/api/services', methods=['GET'])
def get_services():
    """Retorna lista de serviços"""
    services = Service.query.order_by(Service.name).all()
    return jsonify([s.to_dict() for s in services])

@app.route('/api/services', methods=['POST'])
def create_service():
    """Cria um novo serviço"""
    data = request.get_json()
    name = data.get('name', '').strip()
    
    if not name:
        return jsonify({'error': 'Nome do serviço é obrigatório'}), 400
    
    # Verificar se já existe
    if Service.query.filter_by(name=name).first():
        return jsonify({'error': 'Serviço já existe'}), 400
    
    service = Service(name=name)
    db.session.add(service)
    db.session.commit()
    
    return jsonify(service.to_dict()), 201

@app.route('/api/services/<int:service_id>', methods=['DELETE'])
def delete_service(service_id):
    """Deleta um serviço"""
    service = Service.query.get(service_id)
    if not service:
        return jsonify({'error': 'Serviço não encontrado'}), 404
    
    db.session.delete(service)
    db.session.commit()
    
    return jsonify({'message': 'Serviço deletado com sucesso'}), 200

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
